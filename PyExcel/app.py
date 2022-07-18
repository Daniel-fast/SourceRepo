import openpyxl

wb  = openpyxl.load_workbook("transactions.xlsx")


sheet = wb["Sheet1"]

# wb.create_sheet("Sheet2", 1)

# accessing individual cell or a range
cell = sheet["a1"]
value = cell.value
print(wb.sheetnames)
print(value)
print(cell.row)
print(cell.column)
print(cell.coordinate)
print()

#Instead of passing the cell name, we can use coordinates
cell_coordinate = sheet.cell(row=1, column=1)
print("=== Using coordinates ===")
print(cell_coordinate.value)
maxRow = sheet.max_row
maxColumn = sheet.max_column
print("total number of rows:", maxRow, " total number of columns:", maxColumn)
print()

# using coordinates is powerfull when we need to iterate over  the values - with a for statment - for example
print("=== Using for loop ===")
for row in range(1, maxRow + 1):
    for column in range(1, maxColumn + 1):
        cell = sheet.cell(row, column)
        coordinate = cell.coordinate
        value = cell.value
        print("column:", column, " row:", row, "=", coordinate, " value:", value)
print()

# we can get all teh values of a row or column
print("=== Getting values ===")

column = sheet["a"]
print(column) # we get a list of tuples (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>)
print()
# we can get all the columns
columns = sheet["a:c"]
print("Columns (tuples of tuples):", columns) # this will return all the values a through c - actually, we get a list of tuples of tuples
# Columns: ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>), (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, 
# <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>), (<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.C4>))
# to see what a list of tuples inside tuples means, take a look:
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>) - this represents a inner tuple (all values of one column)


print()
# we can use ranges also
sheet["a1:c3"]
sheet[1] # will return all the cells in the first row


sheet.append([1, 2, 3])
wb.save("transaction2.xlsx") # will create a new spreadsheet with a range of values a1:c3 + append

#we have a bunch of methods to use for example:

# sheet.insert_rows
# sheet.insert_cols
# sheet.delete_cols
# sheet.delete_rows
# sheet.add_image
# sheet.calculate_dimension
# sheet.merge_cells
# sheet.add_chart
# sheet.add_table



